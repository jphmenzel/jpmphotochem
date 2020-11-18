# -*- coding: UTF-8 -*-

# The source code below is an algorithm that is introduced in the publication "Predicting Wavelength-Dependent Photochemical Reactivity and Selectivity" 
# by Jan Philipp Menzel, James P. Blinco and Christopher Barner-Kowollik. 
# The code was created by Jan Philipp Menzel.
# Contact by e-mail: j3.menzel@qut.edu.au
# Submitted for publication alongside the manuscript: 1 July 2020
# Purpose of the algorithm: Quantitative prediction of the wavelength-dependent selectivity of two competing photoreactions using monochromatic light.
# Detailed Notes: The algorithm requests manual input, imports uv/vis data from a respective excel file, makes lists for spatial distribution of reactands and products, 
# calculates time dependent development of overall conversion for varied monochromatic wavelengths at requested amount of reactands as well as calculates the required photon count.
# jpmenzelqled.py
print('This algorithm is introduced in the publication: Predicting Wavelength-Dependent Photochemical Reactivity and Selectivity by Jan Philipp Menzel, James P. Blinco and Christopher Barner-Kowollik.')
print('Purpose of the algorithm: Quantitative prediction of the wavelength-dependent selectivity of two competing photoreactions (A and B react with NEM, refer to the above-mentioned publication) using monochromatic light. ')
print("Please ensure that the required excel files can be located by the algorithm.")
import math
import openpyxl
import datetime
# begin input fixed parameters
vsolvent=eval(input('Enter the volume of solvent per vial (e.g. 0.11) [mL]:'))
vsolventinput=vsolvent
vsolventfixed=vsolvent*0.001*0.001 # change to m³
vsolvent=vsolventfixed
targetconversion=97		#set to which conversion simulation should be run for each wavelength, max 99.9
# begin set experimental and simulation parameters
#lambdastart=285						# SET LAMBDA 
#lambdaend=400						#SET LAMBDA
lambdastart=eval(input('Enter the shorter wavelength (e.g. 285) [nm]:'))
lambdaend=eval(input('Enter the longer wavelength (e.g. 390) [nm]:'))
lambdastep=0.5
nst=1
nra=eval(input('Enter the amount of A n(A) (e.g. 0.7) [micromol]:'))
nra=nra*0.001*0.001 # convert to mol
nrb=eval(input('Enter the amount of B n(B) (e.g. 0.7) [micromol]:'))
nrb=nrb*0.001*0.001
nrc=eval(input('Enter the amount of NEM n(N-ethylmaleimide) (e.g. 1.7) [micromol]:'))
nrc=nrc*0.001*0.001 # convert to mol
ncaa=eval(input('Enter the amount of HNBA n(2-hydroxy-5-nitro-benzaldehyde, e.g. 0.0) [micromol]:'))
ncaa=ncaa*0.001*0.001 # convert to mol
nrax=nra
nrbx=nrb
nrcx=nrc
npx=0.000000000000000
ncaax=ncaa
nphinit=0.0000000005000000000	# incident (x=0) number of photons per pulse [mol] 0.5 nmol per pulse
nphinitfixed=nphinit
xtotalsi=vsolvent/(3.141*(0.0055/2)*(0.0055/2))	# pathlength [m]
ndx=eval(input('How many segments should the solution be divided into for simulation (e.g. 100):'))	# CODE ASKS FOR SPATIAL RESOLUTION	
before = datetime.datetime.now()		# measure calculation time from now on
dx=xtotalsi/ndx # thickness of each segment [m]
vsolvseg=vsolvent/ndx # volume of each segment
xraseg=nra/ndx	# mol compound per segment (initially)
xrbseg=nrb/ndx
xrcseg=nrc/ndx
xcaaseg=ncaa/ndx
xapseg=0.00000000000000000000
xbpseg=0.00000000000000000000
# end set experimental and simulation parameters #######
# begin read uv/vis spectra and make molar attenuation coefficient lists: macalist, macblist, macclist, macaplist, macbplist  #######
wb=openpyxl.load_workbook('jpmenzeluvvisqselectivity.xlsx')			# load excel file from home folder # LOAD MOLAR ATTENUATION COEFFICIENTS
sheetra=wb['a']
sheetrb=wb['b']
sheetrc=wb['nem']
sheetap=wb['ap']
sheetbp=wb['bp']
sheetcaa=wb['hnba']
macalist=[]
macblist=[]
macclist=[]
macaplist=[]
macbplist=[]
maccaalist=[]
maclistindex=[]
xn=5				# build maclistindex
specx=1
xs=sheetra.cell(row=xn, column=specx)
xx=xs.value
searchlambda=xx
while searchlambda<lambdastart:
	xn=xn+nst
	xs=sheetra.cell(row=xn, column=specx)
	xx=xs.value
	searchlambda=xx
xn=xn+nst
xs=sheetra.cell(row=xn, column=specx)
xx=xs.value
searchlambda=xx
indexstart=xn
while searchlambda<lambdaend:
	xs=sheetra.cell(row=xn, column=specx)
	xx=xs.value
	maclistindex.append(xx)
	searchlambda=xx
	xn=xn+nst
indexend=xn
xn=5				# build macalist
specy=2
while xn<indexstart:
	xn=xn+nst
xn=xn+nst
while xn<indexend:	
	ys=sheetra.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)	
	macalist.append(y)
	xn=xn+nst
xn=5				# build macblist
specy=2
while xn<indexstart:
	xn=xn+nst
while xn<indexend:
	ys=sheetrb.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)
	macblist.append(y)
	xn=xn+nst
xn=5				# build macclist
specy=2
count=1
while xn<indexstart:
	xn=xn+nst
while xn<indexend:
	ys=sheetrc.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)
	macclist.append(y)
	xn=xn+nst
xn=5				# build macaplist
specy=2
count=1
while xn<indexstart:
	xn=xn+nst
while xn<indexend:
	ys=sheetap.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)
	macaplist.append(y)
	xn=xn+nst
xn=5				# build macbplist
specy=2
count=1
while xn<indexstart:
	xn=xn+nst
while xn<indexend:
	ys=sheetbp.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)
	macbplist.append(y)
	xn=xn+nst
xn=5				# build maccaalist
specy=2
count=1
while xn<indexstart:
	xn=xn+nst
while xn<indexend:
	ys=sheetcaa.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)
	maccaalist.append(y)
	xn=xn+nst
# end read uv/vis spectra and make molar attenuation coefficient lists: macalist, macblist, macclist, macaplist, macbplist)
# Begin build nphinitlist (list of numbers of photons dependent on wavelength impacting sample)
nphinitlist=[]
lambdalist=[]		#list of wavelengths investigated
emissionlist=[]
llcount=lambdastart		#Begin build lambdalist
while llcount<lambdaend:
	lambdalist.append(llcount)
	llcount=llcount+lambdastep		#End build lambdalist
xn=5
while xn<indexstart:
	xn=xn+nst
xn=xn+nst
while xn<indexend:
	nphinitlk=nphinitfixed		#Simulation is run with an equal number of incident photons (before transmittance through glass vial) per laser pulse, irrespective of wavelength 
	nphinitlist.append(nphinitlk)
	xn=xn+nst
nphinitlk=nphinitfixed
nphinitlist.append(nphinitlk)
xn=xn+nst
# begin make empty lists
xralist=[]
xrblist=[]
xrclist=[]
xaplist=[]
xbplist=[]
xcaalist=[]
xlistindex=[]
xperclistindex=[]
tconvalist=[]
tconvblist=[]
tconvlistindex=[]
selectalist=[]	# make selectlists: lists of conversion values of conversion of A and B and wavelength and time of irradiation
selectblist=[]
selectwlist=[]
selecttlist=[]
conralist=[]
conrblist=[]
conrclist=[]
conaplist=[]
conbplist=[]
lambdatconvalist=[]
lambdatconvblist=[]
lambdatconvlistindex=[]
lambdaindex=[]
# end make empty lists
# begin full iteration 
w=lambdastart
we=lambdaend
xw=0			# begin iteration over wavelengths 
# begin determine molar attenuation coefficients at this wavelength
wi=0
ww=maclistindex[wi]
wtarget=w
while ww<wtarget:
	wi=wi+1
	ww=maclistindex[wi]
maca=macalist[wi]	# molar attenuation coefficient at current wavelength for reactand a
macb=macblist[wi]	# molar attenuation coefficient at current wavelength for reactand b
macc=macclist[wi]	# molar attenuation coefficient at current wavelength for reactand c
macap=macaplist[wi]	# molar attenuation coefficient at current wavelength for product ap
macbp=macbplist[wi]	# molar attenuation coefficient at current wavelength for product bp
maccaa=maccaalist[wi] # molar attenuation coefficient at current wavelength for competitive absorber A caa
# end determine molar attenuation coefficients at this wavelength
while w<we:	
	t=0.000000000000000
	coconva=0
	coconvb=0
	conralist=[]
	conrblist=[]
	conrclist=[]
	conaplist=[]
	conbplist=[]
	# begin generate initial state: spatial distribution of species at t=0 along x axis (pathlength of laser beam)
	xinitial=0.00000000000000000
	x=0	
	xralist=[]
	xrblist=[]
	xrclist=[]
	xaplist=[]
	xbplist=[]
	xcaalist=[]
	xlistindex=[]
	xperclistindex=[]
	while x<ndx:
		xlistindex.append(xinitial)
		xperc=(xinitial/xtotalsi)*100.00000000000000000
		xralist.append(xraseg)			#contains reactant [mol] in each slice
		xrblist.append(xrbseg)
		xrclist.append(xrcseg)
		xaplist.append(xapseg)	
		xbplist.append(xbpseg)
		xcaalist.append(xcaaseg)	
		xinitial=xinitial+dx
		x=x+1
	# end generate initial state: spatial distribution of species at t=0 along x axis (pathlength of laser beam)
	# begin determine molar attenuation coefficients at this wavelength
	wi=0
	ww=maclistindex[wi]
	wtarget=w
	while ww<wtarget:
		wi=wi+1
		ww=maclistindex[wi]
	maca=macalist[wi]	# molar attenuation coefficient at current wavelength for reactand a
	macb=macblist[wi]	# molar attenuation coefficient at current wavelength for reactand b
	macc=macclist[wi]	# molar attenuation coefficient at current wavelength for reactand c
	macap=macaplist[wi]	# molar attenuation coefficient at current wavelength for product ap
	macbp=macbplist[wi]	# molar attenuation coefficient at current wavelength for product bp
	maccaa=maccaalist[wi] # molar attenuation coefficient at current wavelength for competitive absorber A caa
	# end determine molar attenuation coefficients at this wavelength
	# begin iteration over laser pulses and generate tconvlist and tconvlistindex
	np=0			#pulse count
	endloopcondition=0
	while endloopcondition<1:		# begin iteration over time
		lamtrigger=0				
		maca=macalist[wi]	# molar attenuation coefficient at current wavelength for reactand a
		macb=macblist[wi]	# molar attenuation coefficient at current wavelength for reactand b
		macc=macclist[wi]	# molar attenuation coefficient at current wavelength for reactand c
		macap=macaplist[wi]	# molar attenuation coefficient at current wavelength for product ap
		macbp=macbplist[wi]	# molar attenuation coefficient at current wavelength for product bp
		maccaa=maccaalist[wi] # molar attenuation coefficient at current wavelength for competitive absorber A caa
		termc=math.exp(((349.11562-(w)))/(21.22769))		#determine glass transmittance at current wavelength
		termk=math.exp((-1)*((((w)-284.58711)**2)/(2*(10.61099)**2)))
		termn=math.exp((-1)*((((w)-417.92409)**2)/(2*(40.68863)**2)))
		termcut=((((270.935-w)**2)**0.5)+w-270.935)/(2*(w-270.935))
		y=(82.91639-(1.91056*termc)-(15.89814*termk)+(2.91878*termn))*termcut
		transmittance=y*0.01 	# convert % values to values between 0 and 1	
		# end determine molar attenuation coefficients at this wavelength
		nphin=nphinitlist[xw]
		nphin=nphin*transmittance
		x=0		# begin go through solution, each segment from x to xnext: LASER PULSE - photon attenuation and reaction: new spatial distribution generated
		xnext=x+1 	
		while x<ndx:
			nphout=nphin*(1.0000000000000000000000000/(10**(((maca*((xralist[x])/(vsolvseg)))+(macb*((xrblist[x])/(vsolvseg)))+(macc*((xrclist[x])/(vsolvseg)))+(macap*((xaplist[x])/(vsolvseg)))+(macbp*((xbplist[x])/(vsolvseg)))+(maccaa*((xcaalist[x])/(vsolvseg))))*dx))) 	
			# attenuation of number of photons by solution in segment
			conccurrentc=(xrclist[x]/vsolvseg)*0.001 # calc current concentration of C and convert to mol L-1
			conccurrent=(xralist[x]/vsolvseg)*0.001 # calc current concentration of A and convert to mol L-1
			terml=math.exp((-1)*((((w)-312)**2)/(2*(12.5)**2)))
			termo=math.exp((-1)*((((w)-294)**2)/(2*(9)**2)))
			termk=math.exp((-1)*((((w)-423.001)**2)/(2*(20)**2)))
			termt=math.exp((-1)*((((w)-397)**2)/(2*(9)**2)))
			termc=math.exp((-1)*75*(conccurrent**0.715))
			termcut=((((w-423.001)**2)**0.5)+423.001-w)/(2*(423.001-w))
			qyacurrent=(0.0272+(0.065*terml)+(0.035*termo)-(0.0272*termk)+(0.01*termt))*2.35*termc*termcut
			nphabstot=nphin-nphout 	# calculation of photons absorbed by solution in total
			# calculation of photons absorbed by A: nphabsa
			nphabsa=nphabstot*((maca*((xralist[x])/vsolvseg))/(((maca*((xralist[x])/vsolvseg))+(macb*((xrblist[x])/vsolvseg))+(macc*((xrclist[x])/vsolvseg))+(macap*((xaplist[x])/vsolvseg))+(macbp*((xbplist[x])/vsolvseg))+(maccaa*((xcaalist[x])/vsolvseg)))))
			nareacted=qyacurrent*nphabsa
			xralist[x]=(xralist[x])-(nareacted)
			xrclist[x]=(xrclist[x])-(nareacted)
			xaplist[x]=(xaplist[x])+(nareacted)		# redefined numbers of compounds in segment after pulse fragment and reaction of A
			# calculation of photons absorbed by B: nphabsb
			nphabsb=nphabstot*((macb*((xrblist[x])/vsolvseg))/(((maca*((xralist[x])/vsolvseg))+(macb*((xrblist[x])/vsolvseg))+(macc*((xrclist[x])/vsolvseg))+(macap*((xaplist[x])/vsolvseg))+(macbp*((xbplist[x])/vsolvseg))+(maccaa*((xcaalist[x])/vsolvseg)))))
			termba=math.exp((-1)*((((w)-288)**2)/(2*(21)**2)))		#terms for calculation of QY of B in dependence of wavelength
			termbb=math.exp((-1)*((((w)-321)**2)/(2*(30)**2)))
			termbc=math.exp((-1)*((((w)-347)**2)/(2*(13)**2)))
			termbd=math.exp((-1)*((((w)-360)**2)/(2*(10)**2)))
			termbe=math.exp((-1)*((((w)-377)**2)/(2*(12)**2)))
			termbf=math.exp((-1)*((((w)-400)**2)/(2*(5)**2)))
			termcutb=((((w-400.999)**2)**0.5)+400.999-w)/(2*(400.999-w))
			qybcurrent=((0.35*termba)+(0.78*termbb)+(0.29*termbc)+(0.08*termbd)+(0.17*termbe)-(0.045*termbf))*termcutb
			nbreacted=qybcurrent*nphabsb			
			xrblist[x]=(xrblist[x])-(nbreacted)
			xrclist[x]=(xrclist[x])-(nbreacted)
			xbplist[x]=(xbplist[x])+(nbreacted)		# redefined numbers of compounds in segment after pulse fragment and reaction of B
			nphin=nphout
			x=x+1
			xnext=xnext+1   	# end go through solution, each segment from x to xnext: LASER PULSE - photon attenuation and reaction: new spatial distribution generated
		w=w
		xw=xw
		# begin calculate current overall conversion, save in tconvlist
		apsum=0.000000000000000000000000000000
		bpsum=0.000000000000000000000000000000
		rasum=0.000000000000000000000000000000
		rbsum=0.000000000000000000000000000000
		rcsum=0.000000000000000000000000000000
		clcx=0
		while clcx<ndx:
			rasum=xralist[clcx]+rasum
			rbsum=xrblist[clcx]+rbsum
			rcsum=xrclist[clcx]+rcsum
			apsum=xaplist[clcx]+apsum
			bpsum=xbplist[clcx]+bpsum
			clcx=clcx+1
		coconva=(apsum/((rasum)+apsum))*100.00000000000000000	# current overall conversion of A
		coconvb=(bpsum/((rbsum)+bpsum))*100.00000000000000000	# current overall conversion of B
		conralist.append(rasum)
		conrclist.append(rcsum)
		conrblist.append(rbsum)
		conaplist.append(apsum)
		conbplist.append(bpsum)
		tconvalist.append(coconva)
		tconvblist.append(coconvb)
		npnext=np+1
		tconvlistindex.append(npnext)	# saved current overall conversion in tconvlist
		# end calculate current overall conversion, save in tconvlist
		endloopcondition=0
		if targetconversion<coconva:
			endloopcondition=1
		if targetconversion<coconvb:
			endloopcondition=1
		# begin go through solution, DARK TIME COMPLETE MIXING - new spatial distribution !
		x=0
		while x<ndx:
			xralist[x]=rasum/ndx
			xrblist[x]=rbsum/ndx
			xrclist[x]=rcsum/ndx
			xaplist[x]=apsum/ndx
			xbplist[x]=bpsum/ndx
			x=x+1	
		x=0			
		# end go through solution, DARK TIME COMPLETE MIXING - new spatial distribution !
		np=np+1
		# end iteration over laser pulses and generate tconvlist and tconvlistindex #######
	lambdatconvalist.append(tconvalist)
	lambdatconvblist.append(tconvblist)
	lambdatconvlistindex.append(tconvlistindex)
	lambdaindex.append(w)	
	selectalist.append(coconva)
	selectblist.append(coconvb)
	selecttlist.append(np)
	selectwlist.append(w)
	w=w+0.5
	xw=xw+1
	wi=wi+1
# end full iteration
# begin save to excel file 
after = datetime.datetime.now()
dt=after-before
wb=openpyxl.load_workbook('jpmenzelqselectivityoutputread.xlsx')
r=1
c=1
sheet=wb['output']
sheet.cell(row=r, column=1).value='wavelength'			# write top left cell
sheet.cell(row=r, column=2).value='p(A) / %'			# write top cell 2
sheet.cell(row=r, column=3).value='p(B) / %'			# write top cell 3
sheet.cell(row=r, column=6).value='number of pulses'	# write top cell 6
# begin write simulation input data and information in column 4 and 5
sheet.cell(row=1, column=4).value='Input data and information'	
sheet.cell(row=1, column=5).value='Selectivity Simulation, full mixing.'	#### MODIFY ?##############
sheet.cell(row=2, column=4).value='Simulation start date and time'
sheet.cell(row=2, column=5).value=before
sheet.cell(row=3, column=4).value='Duration of simulation (hh:mm:ss)'
sheet.cell(row=3, column=5).value=dt
sheet.cell(row=4, column=4).value='Input: V (mL)'
sheet.cell(row=4, column=5).value=vsolventinput
sheet.cell(row=5, column=4).value='Input: n reactant A, PTP (mmol)'
nrap=nra*1000
sheet.cell(row=5, column=5).value=nrap
sheet.cell(row=6, column=4).value='Input: n reactant C, NEM (mmol)'
nrcp=nrc*1000
sheet.cell(row=6, column=5).value=nrcp
sheet.cell(row=7, column=4).value='Input: n reactant B, MPE (mmol)'
nrap=nra*1000
sheet.cell(row=7, column=5).value=nrap
sheet.cell(row=8, column=4).value='Input: n competitive absorber HNBA (mmol)'
ncaap=ncaa*1000
sheet.cell(row=8, column=5).value=ncaap
sheet.cell(row=9, column=4).value='Defined: Incident photons per pulse (before transmittance, mol)'
sheet.cell(row=9, column=5).value=nphinitfixed
sheet.cell(row=10, column=4).value='Input: Number of segments'
sheet.cell(row=10, column=5).value=ndx
sheet.cell(row=11, column=4).value='Wavelength range from'
sheet.cell(row=11, column=5).value=lambdastart
sheet.cell(row=12, column=4).value='Wavelength range to'
sheet.cell(row=12, column=5).value=lambdaend
# end write simulation input data and information in column 4 and 5
tcl=0
r=0
kr=2
while r<len(selectwlist):
	sheet.cell(row=kr, column=1).value=selectwlist[tcl]	# write wavelengths in first column
	sheet.cell(row=kr, column=2).value=selectalist[tcl]	# 
	sheet.cell(row=kr, column=3).value=selectblist[tcl]	# 
	sheet.cell(row=kr, column=6).value=selecttlist[tcl]	# 
	tcl=tcl+1
	r=r+1
	kr=kr+1
wb.save('jpmenzelqselectivitywrite.xlsx')
#print('Conversion data is saved')
print("The simulation is completed, data is saved in excel file jpmenzelqselectivitywrite.xlsx")
print("Time required for calculation was: (h:min:sec)")
print(dt)
# end save to excel file 
