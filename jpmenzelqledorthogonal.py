# -*- coding: UTF-8 -*-

# The source code below is an algorithm that is introduced in the publication "Predicting Wavelength-Dependent Photochemical Reactivity and Selectivity" 
# by Jan Philipp Menzel, James P. Blinco and Christopher Barner-Kowollik. 
# The code was created by Jan Philipp Menzel.
# Contact by e-mail: j3.menzel@qut.edu.au
# Submitted for publication alongside the manuscript: 1 July 2020
# Purpose of the algorithm: Quantitative prediction of wavelength, photon number, time and concentration dependent conversion of photoreaction employing an LED.
# Notes: The below source code is designed to predict conversion of thioether-substituted o-methylbenzaldehyde A with N-ethylmaleiminde NEM (refer to the above-mentioned publication) 
# using LED 2 (emission centered around 343 nm) in the respective 3D-printed photoreactor. 
# Detailed Notes: The algorithm requests manual input, imports uv/vis data from a respective excel file, makes lists for spatial distribution of reactands and products, 
# calculates time dependent development of overall conversion for wavelengths of the respective LED at requested amount of reactands as well as calculates light attenuation maps.
# This algorithm simulates the competing photoreaction between Dodecyl-thioether o-methylbenzaldehyde and a diaryltetrazole (N-Phenyl-p-OMe / C-Phenyl-p-methylester)
# jpmenzelqledorthogonal.py
print('This algorithm is introduced in the publication: Predicting Wavelength-Dependent Photochemical Reactivity and Selectivity by Jan Philipp Menzel, James P. Blinco and Christopher Barner-Kowollik.')
print('Purpose of the algorithm: Quantitative prediction of wavelength, photon number, time and concentration dependent conversion of two competing photoreactions employing an LED. Notes: The below source code is designed to predict conversion of Dodecylthioether-substituted o-methylbenzaldehyde C and Tetrazole D with N-ethylmaleiminde NEM (refer to the above-mentioned publication) using LED 1 (emission centered around 285 nm) in the respective 3D-printed photoreactor. ')
print("Please ensure that the required excel files can be located by the algorithm")
import math
import openpyxl
import datetime
# begin request / set experimental and simulation parameters
errest=eval(input('Press 1 to generate the prediction (best estimate); 2 for the upper limit of the confidence of prediction; or 3 for the lower limit:'))
#askqyguess=eval(input('Use arbitrary quantum yield instead of built-in quantum yield map? Yes: 1, No:0 :'))
askqyguess=1
if askqyguess==1:
	qyguessc=eval(input('Enter value of set apparent quantum yield of dodecylthioether o-methylbenzaldehyde C (arbitrary, wavelength and concentration independent) QY:'))
	qyguessd=eval(input('Enter value of set apparent quantum yield of tetrazole D (arbitrary, wavelength and concentration independent) QY:'))
vsolvent=eval(input('Enter Volume of solvent per vial (eg. 0.25) [mL]:'))
vsolventinput=vsolvent
vsolventfixed=vsolvent*0.001*0.001 # change to m³
vsolvent=vsolventfixed
lambdastart=270						#Set Wavelength range (from)
lambdaend=315						#Set Wavelength range (to)		
lambdastep=0.5						#Wavelength resolution is 0.5 nm
nst=1
nra=eval(input('Enter amount of C n(dodecylthioether o-MBA C, e.g. 0.5) [micromol]:'))
nra=nra*0.001*0.001 # convert to mol
nrd=eval(input('Enter amount of D n(tetrazole D, e.g. 0.5) [micromol]:'))
nrd=nrd*0.001*0.001 # convert to mol
nrc=eval(input('Enter amount of NEM n(N-ethylmaleimide, e.g. 0.6) [micromol]:'))
nrc=nrc*0.001*0.001 # convert to mol
ncaa=eval(input('Enter amount of HNBA n(2-hydroxy-5-nitro-benzaldehyde) [micromol]:'))
ncaa=ncaa*0.001*0.001 # convert to mol
nrax=nra
nrcx=nrc
npx=0.000000000
ncaax=ncaa
irrtime=eval(input('Irradiation time (sec):'))
npulses=irrtime
npulses=npulses*1.00000000000
if errest==3:
	ledpowermw=0.8		# Lower limit of LED power [mW]
elif errest==2:
	ledpowermw=1.2		# Upper limit of LED power [mW]
elif errest==1:
	ledpowermw=1.0		# Best estimate of LED power [mW]
ledpower=ledpowermw*0.001	# convert power in mW to W (J/s) (total energy per second)
xtotalsi=vsolvent/(3.14159265*(0.0055/2)*(0.0055/2))	# pathlength [m]
ndx=eval(input('How many segments should the solution be divided into for simulation (eg. 1000):'))	# SPATIAL RESOLUTION	
before = datetime.datetime.now()		# measure calculation time from now on
dx=xtotalsi/ndx # thickness of each segment [m]
vsolvseg=vsolvent/ndx # volume of each segment
xraseg=nra/ndx	# mol compound A per segment (initially)
xrdseg=nrd/ndx	# mol compound D per segment (initially)
xrcseg=nrc/ndx	# mol compound NEM per segment (initially)
xcaaseg=ncaa/ndx	# mol compound HNBA per segment (initially)
xapseg=0.00000000000	# Initially no product AP present
xdpseg=0.00000000000	# Initially no product DP present
# end request / set experimental and simulation parameters
# begin read uv/vis spectra and make molar attenuation coefficient lists: macalist, macclist, macaplist, maccaalist
wb=openpyxl.load_workbook('jpmenzeluvvisqled.xlsx')			# load excel file from home folder or appropriate folder (MOLAR ATTENUATION COEFFICIENTS)
sheetra=wb['c']
sheetrd=wb['d']
sheetrc=wb['nem']
sheetap=wb['ap']
sheetdp=wb['dp']
sheetcaa=wb['hnba']
macalist=[]
macdlist=[]
macclist=[]
macaplist=[]
macdplist=[]
maccaalist=[]
maclistindex=[]
xn=5				# build maclistindex
specx=1
xs=sheetra.cell(row=xn, column=specx)
xx=xs.value
searchlambda=xx
while searchlambda<lambdastart:
	xn=xn+nst
	xs=sheetra.cell(row=xn, column=specx)
	xx=xs.value
	searchlambda=xx
xn=xn+nst
xs=sheetra.cell(row=xn, column=specx)
xx=xs.value
searchlambda=xx
indexstart=xn
while searchlambda<lambdaend:
	xs=sheetra.cell(row=xn, column=specx)
	xx=xs.value
	maclistindex.append(xx)
	searchlambda=xx
	xn=xn+nst
indexend=xn
xn=5				# build macalist
specy=2
while xn<indexstart:
	xn=xn+nst
xn=xn+nst
while xn<indexend:	
	ys=sheetra.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)	
	macalist.append(y)
	xn=xn+nst
ys=sheetra.cell(row=xn, column=specy)
y=ys.value
if y<0:
	y=0
y=y*0.1 	# convert to SI unit m³/(mol*m)	
macalist.append(y)	
xn=5				# build macdlist
specy=2
while xn<indexstart:
	xn=xn+nst
xn=xn+nst
while xn<indexend:	
	ys=sheetrd.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)	
	macdlist.append(y)
	xn=xn+nst
ys=sheetrd.cell(row=xn, column=specy)
y=ys.value
if y<0:
	y=0
y=y*0.1 	# convert to SI unit m³/(mol*m)	
macdlist.append(y)	
xn=5				# build macdplist
specy=2
while xn<indexstart:
	xn=xn+nst
xn=xn+nst
while xn<indexend:	
	ys=sheetdp.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)	
	macdplist.append(y)
	xn=xn+nst
ys=sheetdp.cell(row=xn, column=specy)
y=ys.value
if y<0:
	y=0
y=y*0.1 	# convert to SI unit m³/(mol*m)	
macdplist.append(y)	
xn=5				# build macclist
specy=2
count=1
while xn<indexstart:
	xn=xn+nst
while xn<indexend:
	ys=sheetrc.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)
	macclist.append(y)
	xn=xn+nst
ys=sheetrc.cell(row=xn, column=specy)
y=ys.value
if y<0:
	y=0
y=y*0.1 	# convert to SI unit m³/(mol*m)
macclist.append(y)	
xn=5				# build macaplist
specy=2
count=1
while xn<indexstart:
	xn=xn+nst
while xn<indexend:
	ys=sheetap.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)
	macaplist.append(y)
	xn=xn+nst
ys=sheetap.cell(row=xn, column=specy)
y=ys.value
if y<0:
	y=0
y=y*0.1 	# convert to SI unit m³/(mol*m)
macaplist.append(y)	
xn=5				# build maccaalist
specy=2
count=1
while xn<indexstart:
	xn=xn+nst
while xn<indexend:
	ys=sheetcaa.cell(row=xn, column=specy)
	y=ys.value
	if y<0:
		y=0
	y=y*0.1 	# convert to SI unit m³/(mol*m)
	maccaalist.append(y)
	xn=xn+nst
ys=sheetcaa.cell(row=xn, column=specy)
y=ys.value
if y<0:
	y=0
y=y*0.1 	# convert to SI unit m³/(mol*m)
maccaalist.append(y)	
# end read uv/vis spectra and make molar attenuation coefficient lists: macalist, macclist, maccaalist, macaplist
# Begin build nphinitlist (list of numbers of photons dependent on wavelength impacting sample)
nphinitlist=[]
lambdalist=[]		#list of wavelengths emitted by LED 1
emissionlist=[]
llcount=lambdastart		#Begin build lambdalist
while llcount<lambdaend:
	lambdalist.append(llcount)
	llcount=llcount+lambdastep		#End build lambdalist
emsum=0
we=lambdastart+0.5
while we<lambdaend:				#calculate sum of values within wavelength range for emission spectrum
	terma=math.exp((-1)*((((we)-284.15613)**2)/(2*(3.86295)**2)))	#determine emission spectrum values at wavelength we
	termd=math.exp((-1)*((((we)-290.19406)**2)/(2*(4.23537)**2)))
	termg=math.exp((-1)*((((we)-294.71877)**2)/(2*(5.79419)**2)))
	termj=math.exp((-1)*((((we)-305.96032)**2)/(2*(21.05534)**2)))
	termm=math.exp((-1)*((((we)-293.96335)**2)/(2*(11.86232)**2)))
	termp=math.exp((-1)*((((we)-276.94418)**2)/(2*(4.04295)**2)))
	em=(0.79979*terma)+(0.23203*termd)+(0.09949*termg)+(0.02333*termj)+(0.08454*termm)+(0.05864*termp)
	emsum=emsum+em
	we=we+0.5
terma=math.exp((-1)*((((we)-284.15613)**2)/(2*(3.86295)**2)))	#determine emission spectrum values at wavelength we
termd=math.exp((-1)*((((we)-290.19406)**2)/(2*(4.23537)**2)))
termg=math.exp((-1)*((((we)-294.71877)**2)/(2*(5.79419)**2)))
termj=math.exp((-1)*((((we)-305.96032)**2)/(2*(21.05534)**2)))
termm=math.exp((-1)*((((we)-293.96335)**2)/(2*(11.86232)**2)))
termp=math.exp((-1)*((((we)-276.94418)**2)/(2*(4.04295)**2)))
em=(0.79979*terma)+(0.23203*termd)+(0.09949*termg)+(0.02333*termj)+(0.08454*termm)+(0.05864*termp)
emsum=emsum+em
#READ EMISSION SPECTRUM AND CALCULATE FRACTIONAL EMITTED PHOTON COUNT
we=lambdastart+0.5
while we<lambdaend:
	terma=math.exp((-1)*((((we)-284.15613)**2)/(2*(3.86295)**2)))	#determine emission spectrum values at wavelength we
	termd=math.exp((-1)*((((we)-290.19406)**2)/(2*(4.23537)**2)))
	termg=math.exp((-1)*((((we)-294.71877)**2)/(2*(5.79419)**2)))
	termj=math.exp((-1)*((((we)-305.96032)**2)/(2*(21.05534)**2)))
	termm=math.exp((-1)*((((we)-293.96335)**2)/(2*(11.86232)**2)))
	termp=math.exp((-1)*((((we)-276.94418)**2)/(2*(4.04295)**2)))
	em=(0.79979*terma)+(0.23203*termd)+(0.09949*termg)+(0.02333*termj)+(0.08454*termm)+(0.05864*termp)
	pcurrent=ledpower*(em/emsum)
	lcurr=we*0.001*0.001*0.001		#change wavelength from nm to m
	nphinitlk=(pcurrent*lcurr)/((6.62607*(10**(-34)))*299792458*(6.02214086*(10**(23))))
	nphinitlist.append(nphinitlk)
	we=we+0.5
terma=math.exp((-1)*((((we)-284.15613)**2)/(2*(3.86295)**2)))	#determine emission spectrum values at wavelength we
termd=math.exp((-1)*((((we)-290.19406)**2)/(2*(4.23537)**2)))
termg=math.exp((-1)*((((we)-294.71877)**2)/(2*(5.79419)**2)))
termj=math.exp((-1)*((((we)-305.96032)**2)/(2*(21.05534)**2)))
termm=math.exp((-1)*((((we)-293.96335)**2)/(2*(11.86232)**2)))
termp=math.exp((-1)*((((we)-276.94418)**2)/(2*(4.04295)**2)))
em=(0.79979*terma)+(0.23203*termd)+(0.09949*termg)+(0.02333*termj)+(0.08454*termm)+(0.05864*termp)
pcurrent=ledpower*(em/emsum)
lcurr=we*0.001*0.001*0.001		#change wavelength from nm to m
nphinitlk=(pcurrent*lcurr)/((6.62607*(10**(-34)))*299792458*(6.02214086*(10**(23))))
nphinitlist.append(nphinitlk)
# begin make empty lists
xralist=[]
xrdlist=[]
xrclist=[]
xaplist=[]
xdplist=[]
xcaalist=[]
xlistindex=[]
xperclistindex=[]
tconvalist=[]
tconvdlist=[]
tconvlistindex=[]
conralist=[]
conrdlist=[]
conrclist=[]
conaplist=[]
condplist=[]
lamtrigger=0	# define for which second of irradiation a light attenuation map shall be recorded
ltone=npulses/3
ltone=int(ltone)
lttwo=npulses/3*2
lttwo=int(lttwo)
ltthree=npulses-1
lamtriggerlist=[0, ltone, lttwo, ltthree]
lamzerolist=[]		# lightattlists at varied wavelengths for defined second - 3D plot
lamonelist=[]
lamtwolist=[]
lamthreelist=[]
lightattlist=[]		# photon count at specific wavelength for defined second - resolved over beampath
lamwlist=[]			# wavelengths in light attenuation map
lambdatconvalist=[]
lambdatconvdlist=[]
lambdatconvlistindex=[]
lambdaindex=[]
# end make empty lists # Required lists are built
# begin full iteration
t=0.000000000000000
# begin generate initial state: spatial distribution of species at t=0 along x axis (pathlength of laser beam)
xinitial=0.00000000000000000
x=0	
while x<ndx:
	xlistindex.append(xinitial)
	xperc=(xinitial/xtotalsi)*100.00000000000000000
	xralist.append(xraseg)			#contains reactant [mol] in each slice
	xrdlist.append(xrdseg)
	xrclist.append(xrcseg)
	xaplist.append(xapseg)	
	xdplist.append(xdpseg)
	xcaalist.append(xcaaseg)	
	xinitial=xinitial+dx
	x=x+1
# end generate initial state: spatial distribution of species at t=0 along x axis (pathlength of laser beam)
# begin determine molar attenuation coefficients at this wavelength
wi=0
w=lambdastart
ww=maclistindex[wi]
wtarget=w
while ww<wtarget:
	wi=wi+1
	ww=maclistindex[wi]
maca=macalist[wi]	# molar attenuation coefficient at current wavelength for reactand a (Dodecylthioether o-MBA C)
macd=macdlist[wi]	# molar attenuation coefficient at current wavelength for reactand d (tetrazole D)
macc=macclist[wi]	# molar attenuation coefficient at current wavelength for reactand c (NEM)
macap=macaplist[wi]	# molar attenuation coefficient at current wavelength for product ap
macdp=macdplist[wi]	# molar attenuation coefficient at current wavelength for product dp
maccaa=maccaalist[wi] # molar attenuation coefficient at current wavelength for competitive absorber (HNBA)
# end determine molar attenuation coefficients at this wavelength
# begin iteration over laser pulses and generate tconvlist and tconvlistindex
np=0			#pulse count
while np<npulses:
	w=lambdastart
	we=lambdaend
	xw=0
	lamtrigger=0
	if np==lamtriggerlist[0]:
		lamtrigger=1
	elif np==lamtriggerlist[1]:
		lamtrigger=2
	elif np==lamtriggerlist[2]:
		lamtrigger=3
	elif np==lamtriggerlist[3]:
		lamtrigger=4
	while w<we: 	# begin iteration over wavelengths
		# begin determine molar attenuation coefficients at this wavelength
		wi=0
		ww=maclistindex[wi]
		wtarget=w
		while ww<wtarget:
			wi=wi+1
			ww=maclistindex[wi]
		maca=macalist[wi]	# molar attenuation coefficient at current wavelength for reactand a
		macd=macdlist[wi]	# molar attenuation coefficient at current wavelength for reactand d
		macc=macclist[wi]	# molar attenuation coefficient at current wavelength for reactand c
		macap=macaplist[wi]	# molar attenuation coefficient at current wavelength for product ap
		macdp=macdplist[wi]	# molar attenuation coefficient at current wavelength for product dp
		maccaa=maccaalist[wi] # molar attenuation coefficient at current wavelength for competitive absorber A caa
		# end determine molar attenuation coefficients at this wavelength
		termc=math.exp(((349.11562-(w)))/(21.22769))		#determine glass transmittance at current wavelength
		termk=math.exp((-1)*((((w)-284.58711)**2)/(2*(10.61099)**2)))
		termn=math.exp((-1)*((((w)-417.92409)**2)/(2*(40.68863)**2)))
		termcut=((((270.935-w)**2)**0.5)+w-270.935)/(2*(w-270.935))
		y=(82.91639-(1.91056*termc)-(15.89814*termk)+(2.91878*termn))*termcut
		y=y*0.01 	# convert % values to values between 0 and 1	
		if errest==3:		# accuracy of measured transmittance: T_lambda +- 3.2% 
			y=y-(y*3.2/100)
		elif errest==2:
			y=y+(y*3.2/100)
		elif errest==1:
			y=y														
		transmittance=y	# glass transmittance at current wavelength
		nphin=nphinitlist[xw]
		nphin=nphin*transmittance
		x=0		
		xnext=x+1 	
		while x<ndx:		# begin go through solution, each segment from x to xnext: Each second: photon attenuation and reaction: new spatial distribution generated
			nphout=nphin*(1.0000000000000000000000000/(10**(((maca*((xralist[x])/(vsolvseg)))+(macd*((xrdlist[x])/(vsolvseg)))+(macc*((xrclist[x])/(vsolvseg)))+(macap*((xaplist[x])/(vsolvseg)))+(macdp*((xdplist[x])/(vsolvseg)))+(maccaa*((xcaalist[x])/(vsolvseg))))*dx))) 	
			# attenuation of number of photons by solution in segment
			conccurrentc=(xrclist[x]/vsolvseg)*0.001 # calc current concentration of C and convert to mol L-1
			conccurrent=(xralist[x]/vsolvseg)*0.001 # calc current concentration of A and convert to mol L-1
			terml=math.exp((-1)*((((w)-312)**2)/(2*(12.5)**2)))
			termo=math.exp((-1)*((((w)-294)**2)/(2*(9)**2)))
			termk=math.exp((-1)*((((w)-423.001)**2)/(2*(20)**2)))
			termt=math.exp((-1)*((((w)-397)**2)/(2*(9)**2)))
			termc=math.exp((-1)*75*(conccurrent**0.715))
			termcut=((((w-423.001)**2)**0.5)+423.001-w)/(2*(423.001-w))
			qyacurrent=(0.0272+(0.065*terml)+(0.035*termo)-(0.0272*termk)+(0.01*termt))*2.35*termc*termcut
			if askqyguess==1:
				qyacurrent=qyguessc
				qydcurrent=qyguessd
			if errest==3:
				qyacurrent=qyacurrent-(qyacurrent*10/100)
				qydcurrent=qydcurrent+(qydcurrent*10/100)
			elif errest==2:
				qyacurrent=qyacurrent+(qyacurrent*10/100)
				qydcurrent=qydcurrent-(qydcurrent*10/100)
			elif errest==1:
				qyacurrent=qyacurrent
			nphabstot=nphin-nphout 	# calculation of photons absorbed by solution in total
			# calculation of photons absorbed by A: nphabsa
			nphabsa=nphabstot*((maca*((xralist[x])/vsolvseg))/(((maca*((xralist[x])/vsolvseg))+(macd*((xrdlist[x])/vsolvseg))+(macc*((xrclist[x])/vsolvseg))+(macap*((xaplist[x])/vsolvseg))+(macdp*((xdplist[x])/vsolvseg))+(maccaa*((xcaalist[x])/vsolvseg)))))
			nphabsd=nphabstot*((macd*((xrdlist[x])/vsolvseg))/(((maca*((xralist[x])/vsolvseg))+(macd*((xrdlist[x])/vsolvseg))+(macc*((xrclist[x])/vsolvseg))+(macap*((xaplist[x])/vsolvseg))+(macdp*((xdplist[x])/vsolvseg))+(maccaa*((xcaalist[x])/vsolvseg)))))
			nareacted=qyacurrent*nphabsa
			ndreacted=qydcurrent*nphabsd
			xralist[x]=(xralist[x])-(nareacted)
			xrdlist[x]=(xrdlist[x])-(ndreacted)
			xrclist[x]=(xrclist[x])-(nareacted)
			xaplist[x]=(xaplist[x])+(nareacted)	
			xdplist[x]=(xdplist[x])+(ndreacted)	# redefined numbers of compounds in segment after pulse fragment and reaction of A
			if lamtrigger>0:
				lightattlist.append(nphin)
			nphin=nphout
			x=x+1
			xnext=xnext+1   	# end go through solution, each segment from x to xnext: LASER PULSE - photon attenuation and reaction: new spatial distribution generated
		if lamtrigger==1:		# extract light attenuation map
			lamzerolist.append(lightattlist)
			lightattlist=[]
			lamwlist.append(w)
		if lamtrigger==2:
			lamonelist.append(lightattlist)
			lightattlist=[]
		if lamtrigger==3:
			lamtwolist.append(lightattlist)
			lightattlist=[]
		if lamtrigger==4:
			lamthreelist.append(lightattlist)
			lightattlist=[]
		w=w+0.5
		xw=xw+1
	# begin calculate current overall conversion, save in tconvlist
	apsum=0.00000000000000000
	rasum=0.00000000000000000
	dpsum=0.00000000000000000
	rdsum=0.00000000000000000
	rcsum=0.00000000000000000
	clcx=0
	while clcx<ndx:
		rasum=xralist[clcx]+rasum
		rdsum=xrdlist[clcx]+rdsum
		rcsum=xrclist[clcx]+rcsum
		apsum=xaplist[clcx]+apsum
		dpsum=xdplist[clcx]+dpsum
		clcx=clcx+1
	coconva=(apsum/((rasum)+apsum))*100.00000000000000000	# current overall conversion of A to AP
	coconvd=(dpsum/((rdsum)+dpsum))*100.00000000000000000	# current overall conversion of A to AP
	conralist.append(rasum)
	conrdlist.append(rdsum)
	conrclist.append(rcsum)
	conaplist.append(apsum)
	condplist.append(dpsum)
	tconvalist.append(coconva)
	tconvdlist.append(coconvd)
	npnext=np+1
	tconvlistindex.append(npnext)	# saved current overall conversion in tconvlist
	# end calculate current overall conversion, save in tconvlist
	# begin go through solution, DARK TIME COMPLETE MIXING - new spatial distribution
	x=0
	while x<ndx:
		xralist[x]=rasum/ndx
		xrdlist[x]=rdsum/ndx
		xrclist[x]=rcsum/ndx
		xaplist[x]=apsum/ndx
		xdplist[x]=dpsum/ndx
		x=x+1	
	x=0			
	# end go through solution, DARK TIME COMPLETE MIXING - new spatial distribution
	np=np+1
	# end iteration over laser pulses and generate tconvlist and tconvlistindex #######
lambdatconvalist.append(tconvalist)
lambdatconvdlist.append(tconvdlist)
lambdatconvlistindex.append(tconvlistindex)
lambdaindex.append(w)	
# end full iteration
# begin save to excel file
after = datetime.datetime.now()
dt=after-before
wb=openpyxl.load_workbook('jpmenzelqledoutputread.xlsx')
sheet=wb['output']
sheet.cell(row=1, column=1).value='t / sec'		# write top left cell
sheet.cell(row=1, column=2).value='p(Dod-S-o-MBA C) / %'		# write top cell 2
sheet.cell(row=1, column=5).value='p(Tetrazole D) / %'		# write top cell 2
# begin write simulation input data and information in column 3 and 4
sheet.cell(row=1, column=3).value='Input data and information'	
sheet.cell(row=1, column=4).value='Simulation with 285 nm LED, full mixing.'	
sheet.cell(row=2, column=3).value='Simulation start date and time'
sheet.cell(row=2, column=4).value=before
sheet.cell(row=3, column=3).value='Duration of simulation (hh:mm:ss)'
sheet.cell(row=3, column=4).value=dt
sheet.cell(row=4, column=3).value='Input: V (mL)'
sheet.cell(row=4, column=4).value=vsolventinput
sheet.cell(row=5, column=3).value='Input: n reactant C (mmol)'
nrap=nra*1000
sheet.cell(row=5, column=4).value=nrap
sheet.cell(row=6, column=3).value='Input: n reactant NEM (mmol)'
nrcp=nrc*1000
sheet.cell(row=6, column=4).value=nrcp
sheet.cell(row=7, column=3).value='Input: Irradiation time (s)'
sheet.cell(row=7, column=4).value=irrtime
sheet.cell(row=8, column=3).value='Input: Power of LED (mW)'
sheet.cell(row=8, column=4).value=ledpowermw
sheet.cell(row=9, column=3).value='Input: Number of segments'
sheet.cell(row=9, column=4).value=ndx
sheet.cell(row=10, column=3).value='Input: n competitive absorber A (mmol)'
ncaap=ncaa*1000
sheet.cell(row=10, column=4).value=ncaap
sheet.cell(row=11, column=3).value='Input: n tetrazole D (mmol)'
nrdp=nrd*1000
sheet.cell(row=11, column=4).value=nrdp
# end write simulation input data and information in column 3 and 4
tcl=0
r=0
kr=2
while r<len(tconvlistindex):
	sheet.cell(row=kr, column=1).value=tconvlistindex[tcl]	# write irradiation time in seconds in first column
	tcl=tcl+1
	r=r+1
	kr=kr+1
ltcl=0
while ltcl<len(lambdatconvalist):
	insertlist=lambdatconvalist[ltcl]
	insertdlist=lambdatconvdlist[ltcl]	
	r=0
	kr=2
	while r<len(insertlist):
		sheet.cell(row=kr, column=2).value=insertlist[r]	# write conversion data A to AP in second column
		sheet.cell(row=kr, column=5).value=insertdlist[r]	# write conversion data D to DP in fifth column
		r=r+1
		kr=kr+1
	ltcl=ltcl+1
#wb.save('jpmenzelqledwrite.xlsx')
print('Conversion data is saved, saving of data for light attenuation maps is in progress ...')
#wb=openpyxl.load_workbook('jpmenzelqledwrite.xlsx')			# load excel file from home folder 
# begin write light attenuation maps
sheet=wb['lamzero']
sheet.cell(row=1, column=1).value='n_p (mol) vs. wavelengths and segments, map 0'		# write top left cell
lamwi=0
lamwx=2
while lamwi<(len(lamwlist)):
	sheet.cell(row=1, column=lamwx).value=lamwlist[lamwi]		# write wavelengths for light attenuation map
	lamwi=lamwi+1
	lamwx=lamwx+1
ppathl=100/ndx
lamdx=2
lamdi=0
while lamdi<ndx:
	ppathlx=ppathl*(lamdx-1)
	sheet.cell(row=lamdx, column=1).value=ppathlx		# write penetrated pathlength for light attenuation map
	lamdx=lamdx+1
	lamdi=lamdi+1
lamwi=0
lamwx=2
lamdi=0
lamdx=2
while lamwi<(len(lamwlist)):
	lamdi=0
	lamdx=2
	while lamdi<ndx:
		sheet.cell(row=lamdx, column=lamwx).value=lamzerolist[lamwi][lamdi]
		lamdi=lamdi+1
		lamdx=lamdx+1
	lamwi=lamwi+1
	lamwx=lamwx+1
#wb.save('jpmenzelqledwrite.xlsx')
#wb=openpyxl.load_workbook('jpmenzelqledwrite.xlsx')			# load excel file from home folder
sheet=wb['lamone']
sheet.cell(row=1, column=1).value='n_p (mol) vs. wavelengths and segments, map 1'		# write top left cell
lamwi=0
lamwx=2
while lamwi<(len(lamwlist)):
	sheet.cell(row=1, column=lamwx).value=lamwlist[lamwi]		# write wavelengths for light attenuation map
	lamwi=lamwi+1
	lamwx=lamwx+1
ppathl=100/ndx
lamdx=2
lamdi=0
while lamdi<ndx:
	ppathlx=ppathl*(lamdx-1)
	sheet.cell(row=lamdx, column=1).value=ppathlx		# write penetrated pathlength for light attenuation map
	lamdx=lamdx+1
	lamdi=lamdi+1
lamwi=0
lamwx=2
lamdi=0
lamdx=2
while lamwi<(len(lamwlist)):
	lamdi=0
	lamdx=2
	while lamdi<ndx:
		sheet.cell(row=lamdx, column=lamwx).value=lamonelist[lamwi][lamdi]
		lamdi=lamdi+1
		lamdx=lamdx+1
	lamwi=lamwi+1
	lamwx=lamwx+1
sheet=wb['lamtwo']
sheet.cell(row=1, column=1).value='n_p (mol) vs. wavelengths and segments, map 2'		# write top left cell
lamwi=0
lamwx=2
while lamwi<(len(lamwlist)):
	sheet.cell(row=1, column=lamwx).value=lamwlist[lamwi]		# write wavelengths for light attenuation map
	lamwi=lamwi+1
	lamwx=lamwx+1
ppathl=100/ndx
lamdx=2
lamdi=0
while lamdi<ndx:
	ppathlx=ppathl*(lamdx-1)
	sheet.cell(row=lamdx, column=1).value=ppathlx		# write penetrated pathlength for light attenuation map
	lamdx=lamdx+1
	lamdi=lamdi+1
lamwi=0
lamwx=2
lamdi=0
lamdx=2
while lamwi<(len(lamwlist)):
	lamdi=0
	lamdx=2
	while lamdi<ndx:
		sheet.cell(row=lamdx, column=lamwx).value=lamtwolist[lamwi][lamdi]
		lamdi=lamdi+1
		lamdx=lamdx+1
	lamwi=lamwi+1
	lamwx=lamwx+1
sheet=wb['lamthree']
sheet.cell(row=1, column=1).value='n_p (mol) vs. wavelengths and segments, map 3'		# write top left cell
lamwi=0
lamwx=2
while lamwi<(len(lamwlist)):
	sheet.cell(row=1, column=lamwx).value=lamwlist[lamwi]		# write wavelengths for light attenuation map
	lamwi=lamwi+1
	lamwx=lamwx+1
ppathl=100/ndx
lamdx=2
lamdi=0
while lamdi<ndx:
	ppathlx=ppathl*(lamdx-1)
	sheet.cell(row=lamdx, column=1).value=ppathlx		# write penetrated pathlength for light attenuation map
	lamdx=lamdx+1
	lamdi=lamdi+1
lamwi=0
lamwx=2
lamdi=0
lamdx=2
while lamwi<(len(lamwlist)):
	lamdi=0
	lamdx=2
	while lamdi<ndx:
		sheet.cell(row=lamdx, column=lamwx).value=lamthreelist[lamwi][lamdi]
		lamdi=lamdi+1
		lamdx=lamdx+1
	lamwi=lamwi+1
	lamwx=lamwx+1
wb.save('jpmenzelqledwrite.xlsx')
# end write light attenuation maps
readrow=npulses+1
wb=openpyxl.load_workbook('jpmenzelqledwrite.xlsx')			# load excel file from home folder 
sheetra=wb['output']
xs=sheetra.cell(row=readrow, column=1)
axs=sheetra.cell(row=readrow, column=2)
dxs=sheetra.cell(row=readrow, column=5)
xx=xs.value
axx=axs.value
dxx=dxs.value
print("After %d seconds of irradiation:" % xx)
print("Predicted conversion of C to CP is %.3f %%" % axx )
print("Predicted conversion of D to DP is %.3f %%" % dxx )
print("The simulation is completed, data is saved in excel file jpmenzelqledwrite.xlsx")
print("Time required for calculation was: (h:min:sec)")
print(dt)
# end save to excel file
